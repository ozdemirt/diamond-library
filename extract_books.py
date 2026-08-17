import os
import zipfile
import csv
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_books(books_dir, output_csv="kitaplar_listesi.csv", output_xlsx="kitaplar_listesi.xlsx"):
    """
    Books klasöründeki çıkarılmış kitapları ve ZIP arşivlerinin
    içindeki tüm kitap isimlerini, formatlarını ve lokasyonlarını CSV ve Excel (.xlsx) dosyasına çıkarır.
    """
    target_dir = os.path.abspath(books_dir)
    csv_path = os.path.abspath(output_csv)
    xlsx_path = os.path.abspath(output_xlsx)

    print("=" * 75)
    print("📚 KİTAP TARAMASI VE EXCEL / CSV OLUŞTURMA BAŞLATILDI")
    print(f"📁 Hedef Klasör : {target_dir}")
    print(f"📄 CSV Dosyası  : {csv_path}")
    print(f"📊 Excel Dosyası: {xlsx_path}")
    print("=" * 75 + "\n")

    books = []
    supported_exts = {'.pdf', '.epub', '.mobi', '.txt', '.doc', '.docx'}
    start_time = time.time()

    for root, dirs, files in os.walk(target_dir):
        for file in files:
            full_path = os.path.join(root, file)
            ext = os.path.splitext(file)[1].lower()

            # Doğrudan klasördeki kitaplar
            if ext in supported_exts:
                try:
                    stat = os.stat(full_path)
                    size_mb = round(stat.st_size / (1024 * 1024), 2)
                except Exception:
                    size_mb = 0.0

                books.append({
                    "Kitap Adı": file,
                    "Format": ext.upper().replace('.', ''),
                    "Kaynak / Durum": "Klasörde Çıkarılmış",
                    "Tam Lokasyon": full_path,
                    "Bulunduğu Klasör": root,
                    "Boyut (MB)": size_mb
                })

            # ZIP arşivleri içindeki kitaplar
            elif ext == '.zip':
                try:
                    with zipfile.ZipFile(full_path, 'r') as z:
                        for zinfo in z.infolist():
                            if not zinfo.is_dir():
                                z_ext = os.path.splitext(zinfo.filename)[1].lower()
                                if z_ext in supported_exts:
                                    filename = os.path.basename(zinfo.filename)
                                    size_mb = round(zinfo.file_size / (1024 * 1024), 2)
                                    books.append({
                                        "Kitap Adı": filename,
                                        "Format": z_ext.upper().replace('.', ''),
                                        "Kaynak / Durum": f"ZIP İçinde ({file})",
                                        "Tam Lokasyon": f"{full_path} -> {zinfo.filename}",
                                        "Bulunduğu Klasör": os.path.dirname(zinfo.filename) or file,
                                        "Boyut (MB)": size_mb
                                    })
                except Exception as e:
                    print(f"⚠️ ZIP okuma hatası ({file}): {e}")

    elapsed = time.time() - start_time

    # 1. CSV Kaydı (Excel ile tam uyum için utf-8-sig)
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        fieldnames = ['Kitap Adı', 'Format', 'Kaynak / Durum', 'Tam Lokasyon', 'Bulunduğu Klasör', 'Boyut (MB)']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(books)

    # 2. Şık Excel (.xlsx) Oluşturma
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kitap Listesi"

    headers = ['Kitap Adı', 'Format', 'Kaynak / Durum', 'Tam Lokasyon', 'Bulunduğu Klasör', 'Boyut (MB)']
    ws.append(headers)

    for item in books:
        ws.append([
            item["Kitap Adı"],
            item["Format"],
            item["Kaynak / Durum"],
            item["Tam Lokasyon"],
            item["Bulunduğu Klasör"],
            item["Boyut (MB)"]
        ])

    # Excel Stilleri
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')

    zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Başlık Satırı Biçimlendirme
    ws.row_dimensions[1].height = 28
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Veri Satırları Biçimlendirme
    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 20
        is_even = (row_num % 2 == 0)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if col_num in [2, 6]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Filtreleme Özelliği Ekleme
    ws.auto_filter.ref = ws.dimensions

    # Sütun Genişliklerini Otomatik Ayarlama
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 80)

    wb.save(xlsx_path)

    print("\n" + "=" * 75)
    print("✅ İŞLEM TAMAMLANDI!")
    print(f"📚 Bulunan Toplam Kitap Sayısı : {len(books)}")
    print(f"⏱️ Geçen Süre                    : {elapsed:.2f} saniye")
    print(f"📄 Oluşturulan CSV Dosyası       : {csv_path}")
    print(f"📊 Oluşturulan Excel Dosyası     : {xlsx_path}")
    print("=" * 75)

if __name__ == "__main__":
    books_directory = r"C:\Users\default.LAPTOP-OBA1HVU9\.antigravity-ide\Books"
    extract_books(books_directory)
