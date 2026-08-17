import os
import zipfile
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_non_pdf_books(target_dir=r"C:\Users\default.LAPTOP-OBA1HVU9\.antigravity-ide\Books",
                          csv_path="pdf_olmayan_kitaplar.csv",
                          xlsx_path="pdf_olmayan_kitaplar.xlsx"):
    """
    Books klasöründeki PDF OLMAYAN tüm kitap ve dokümanları (EPUB, DOCX, PPTX vb.)
    tarlar ve hem CSV hem de Excel (.xlsx) olarak kaydeder.
    """
    non_pdf_books = []
    supported_non_pdf_exts = {'.epub', '.mobi', '.txt', '.doc', '.docx', '.ppt', '.pptx', '.html'}

    for root, dirs, files in os.walk(target_dir):
        for f in files:
            full_path = os.path.join(root, f)
            ext = os.path.splitext(f)[1].lower()

            if ext in supported_non_pdf_exts:
                try:
                    size_mb = round(os.path.getsize(full_path) / (1024 * 1024), 2)
                except Exception:
                    size_mb = 0.0
                non_pdf_books.append({
                    'Kitap Adı': f,
                    'Format': ext.upper().replace('.', ''),
                    'Kaynak / Durum': 'Klasörde Çıkarılmış',
                    'Bulunduğu Klasör': root,
                    'Tam Dosya Yolu': full_path,
                    'Boyut (MB)': size_mb
                })
            elif ext == '.zip':
                try:
                    with zipfile.ZipFile(full_path, 'r') as z:
                        for zinfo in z.infolist():
                            if not zinfo.is_dir():
                                z_ext = os.path.splitext(zinfo.filename)[1].lower()
                                if z_ext in supported_non_pdf_exts:
                                    filename = os.path.basename(zinfo.filename)
                                    if filename:
                                        size_mb = round(zinfo.file_size / (1024 * 1024), 2)
                                        non_pdf_books.append({
                                            'Kitap Adı': filename,
                                            'Format': z_ext.upper().replace('.', ''),
                                            'Kaynak / Durum': f'ZIP Arşivi İçinde ({f})',
                                            'Bulunduğu Klasör': os.path.dirname(zinfo.filename) or f,
                                            'Tam Dosya Yolu': f'{full_path} -> {zinfo.filename}',
                                            'Boyut (MB)': size_mb
                                        })
                except Exception as e:
                    print(f"⚠️ ZIP okuma hatası ({f}): {e}")

    # CSV Kaydı
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        fieldnames = ['Kitap Adı', 'Format', 'Kaynak / Durum', 'Bulunduğu Klasör', 'Tam Dosya Yolu', 'Boyut (MB)']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(non_pdf_books)

    # Excel Kaydı
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PDF Olmayan Kitaplar'

    headers = ['Kitap Adı', 'Format', 'Kaynak / Durum', 'Bulunduğu Klasör', 'Tam Dosya Yolu', 'Boyut (MB)']
    ws.append(headers)

    for item in non_pdf_books:
        ws.append([
            item['Kitap Adı'],
            item['Format'],
            item['Kaynak / Durum'],
            item['Bulunduğu Klasör'],
            item['Tam Dosya Yolu'],
            item['Boyut (MB)']
        ])

    header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')

    zebra_fill = PatternFill(start_color='FDF2F2', end_color='FDF2F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    ws.row_dimensions[1].height = 28
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20
        is_even = (r % 2 == 0)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if c in [2, 6]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 75)

    wb.save(xlsx_path)

    print(f"✅ Toplam {len(non_pdf_books)} adet PDF olmayan kitap/doküman bulundu.")

if __name__ == "__main__":
    extract_non_pdf_books()
