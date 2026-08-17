import os
import re
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def detect_author(filename, folder, full_path):
    path = full_path.replace('\\', '/')
    name = filename.strip()
    name_lower = name.lower()

    if any(k in path for k in ['/Suat Yıldırım/', '/Suat Yildirim/', 'Suat Yildirim', 'Suat Yıldırım']):
        return 'Suat Yıldırım'
    if any(k in path for k in ['/Hekimoğlu İsmail/', 'Hekimoğlu', 'Hekimoglu']):
        return 'Hekimoğlu İsmail'
    if any(k in path for k in ['/MFG/', '/PRL-English/', 'Fethullah', 'Gulen', 'Fasildan-Fasila', 'Kalbin-Zumrut', 'Inancin-Golgesinde', 'Cag-ve-Nesil']):
        return 'M. Fethullah Gülen'
    if any(k in path for k in ['Necmettin-Sahiner', 'Necmeddin-Şahiner', 'Sahiner', 'Necmeddin Şahiner']):
        return 'Necmeddin Şahiner'
    if ('Davut' in name and ('Aydüz' in name or 'Ayduz' in name)) or 'Davut Aydüz' in path:
        return 'Davut Aydüz'
    if 'Vehbe Zuhayli' in name or 'Zuhayli' in name or 'vehbe-zuhayli' in name_lower:
        return 'Vehbe Zuhayli'
    if 'Ali Unal' in name or 'ali-unal' in name_lower or 'Ali Ünal' in name:
        return 'Ali Ünal'
    if 'Aytunc Altindal' in name or 'aytunc-altindal' in name_lower or 'Aytunç Altındal' in name:
        return 'Aytunç Altındal'
    if 'Reşid Haylamaz' in path or 'Haylamaz' in path or 'haylamaz' in name_lower:
        return 'Reşit Haylamaz'
    if 'Mehmet Akar' in path or 'mehmet-akar' in name_lower:
        return 'Mehmet Akar'
    if 'Umberto Eco' in name or 'umberto-eco' in name_lower:
        return 'Umberto Eco'
    if 'İbrahim Canan' in name or 'ibrahim-canan' in name_lower:
        return 'İbrahim Canan'
    if 'rabia-kandira' in name_lower:
        return 'Rabia Kandıra'
    if 'osman-kaplan' in name_lower:
        return 'Osman Kaplan'
    if 'safak-demir' in name_lower:
        return 'Şafak Demir'
    if 'h-ibrahim-cayirli' in name_lower or 'ibrahim-cayirli' in name_lower:
        return 'H. İbrahim Çayırlı'
    if 'aysel-ertan' in name_lower:
        return 'Aysel Ertan'
    if 'ayhan-tekines' in name_lower or 'Ayhan Tekineş' in path:
        return 'Ayhan Tekineş'
    if 'Arif Sarsılmaz' in path or 'sarsilmaz' in name_lower:
        return 'Arif Sarsılmaz'
    if 'sayit-kocer' in name_lower:
        return 'Sayıt Koçer'
    if 'Feridun M. Emecan' in name or 'Feridun-M-Emecan' in name:
        return 'Feridun M. Emecan'
    if 'Rainer Hermann' in name:
        return 'Rainer Hermann'
    if 'Erkam Tufan Aytav' in name:
        return 'Erkam Tufan Aytav'
    if 'Kadir Ozkose' in name or 'kadir-ozkose' in name_lower:
        return 'Kadir Özköse'

    match = re.match(r'^([A-ZÇĞİÖŞÜ][a-zçğıöşü]+(?:\s+[A-ZÇĞİÖŞÜ][a-zçğıöşü]+)+)\s*-\s*', name)
    if match:
        cand = match.group(1).strip()
        if len(cand) > 4 and cand not in ['Mesnevi Hikayelerinden', 'Final sample', 'Lab and', 'Kuresel Dunyada', 'Untitled document']:
            return cand

    return 'Çeşitli / Konu Bazlı'

def process_books_by_author(csv_input="kitaplar_listesi.csv", csv_output="kitaplar_yazarlarina_gore.csv", xlsx_output="kitaplar_yazarlarina_gore.xlsx"):
    all_rows = []
    with open(csv_input, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            author = detect_author(row['Kitap Adı'], row['Bulunduğu Klasör'], row['Tam Dosya Yolu'])
            try:
                size_mb = float(row['Boyut (MB)'])
            except Exception:
                size_mb = 0.0

            all_rows.append({
                'Yazar': author,
                'Kitap Adı': row['Kitap Adı'],
                'Format': row['Dosya Tipi'],
                'Kaynak / Durum': row['Kaynak / Durum'],
                'Bulunduğu Klasör': row['Bulunduğu Klasör'],
                'Tam Dosya Yolu': row['Tam Dosya Yolu'],
                'Boyut (MB)': size_mb
            })

    all_rows.sort(key=lambda x: (x['Yazar'], x['Kitap Adı'].lower()))

    # Write CSV
    with open(csv_output, 'w', newline='', encoding='utf-8-sig') as f:
        fieldnames = ['Yazar', 'Kitap Adı', 'Format', 'Kaynak / Durum', 'Bulunduğu Klasör', 'Tam Dosya Yolu', 'Boyut (MB)']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    # Write XLSX
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    summary_fill = PatternFill(start_color='203764', end_color='203764', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')

    zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = 'Yazar Özeti'

    ws_summary.append(['Yazar Adı', 'Kitap Sayısı', 'Toplam Boyut (MB)'])
    author_stats = {}
    for r in all_rows:
        auth = r['Yazar']
        if auth not in author_stats:
            author_stats[auth] = {'count': 0, 'size': 0.0}
        author_stats[auth]['count'] += 1
        author_stats[auth]['size'] += r['Boyut (MB)']

    for auth, stat in sorted(author_stats.items(), key=lambda x: x[1]['count'], reverse=True):
        ws_summary.append([auth, stat['count'], round(stat['size'], 2)])

    ws_summary.row_dimensions[1].height = 28
    for c in range(1, 4):
        cell = ws_summary.cell(row=1, column=c)
        cell.fill = summary_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in range(2, ws_summary.max_row + 1):
        ws_summary.row_dimensions[r].height = 20
        is_even = (r % 2 == 0)
        for c in range(1, 4):
            cell = ws_summary.cell(row=r, column=c)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if c in [2, 3]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws_summary.auto_filter.ref = ws_summary.dimensions
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Main Sheet
    ws_main = wb.create_sheet(title='Tüm Kitaplar (Yazarlı)')
    headers = ['Yazar', 'Kitap Adı', 'Format', 'Kaynak / Durum', 'Bulunduğu Klasör', 'Tam Dosya Yolu', 'Boyut (MB)']
    ws_main.append(headers)

    for r in all_rows:
        ws_main.append([
            r['Yazar'], r['Kitap Adı'], r['Format'], r['Kaynak / Durum'],
            r['Bulunduğu Klasör'], r['Tam Dosya Yolu'], r['Boyut (MB)']
        ])

    ws_main.row_dimensions[1].height = 28
    for c in range(1, len(headers) + 1):
        cell = ws_main.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in range(2, ws_main.max_row + 1):
        ws_main.row_dimensions[r].height = 20
        is_even = (r % 2 == 0)
        for c in range(1, len(headers) + 1):
            cell = ws_main.cell(row=r, column=c)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
            if c in [3, 7]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws_main.auto_filter.ref = ws_main.dimensions
    for col in ws_main.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_main.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 70)

    wb.save(xlsx_output)
    print("İşlem başarıyla tamamlandı!")

if __name__ == "__main__":
    process_books_by_author()
